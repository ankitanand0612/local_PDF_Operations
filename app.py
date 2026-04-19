import os
import io
import zipfile
import threading
import webbrowser
from flask import Flask, request, send_file, render_template_string, jsonify
import pikepdf
import pdfplumber
import openpyxl

app = Flask(__name__)

HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1.0"/>
  <title>PDF Tools</title>
  <style>
    *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

    body {
      display: flex;
      height: 100vh;
      background: #0f0f13;
      font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
      color: #e0e0e0;
      overflow: hidden;
    }

    /* ── Sidebar ── */
    .sidebar {
      width: 230px;
      min-width: 230px;
      background: #13131c;
      border-right: 1px solid #22222f;
      display: flex;
      flex-direction: column;
      padding: 28px 16px 24px;
      gap: 6px;
    }

    .brand {
      display: flex;
      align-items: center;
      gap: 11px;
      padding: 0 8px;
      margin-bottom: 28px;
    }

    .brand-icon {
      width: 36px; height: 36px;
      background: linear-gradient(135deg, #6c63ff, #a855f7);
      border-radius: 9px;
      display: flex; align-items: center; justify-content: center;
      font-size: 18px; flex-shrink: 0;
    }

    .brand h1 { font-size: 17px; font-weight: 700; color: #fff; }
    .brand p  { font-size: 11px; color: #555; margin-top: 1px; }

    .nav-label {
      font-size: 10px;
      font-weight: 600;
      letter-spacing: .08em;
      text-transform: uppercase;
      color: #444;
      padding: 0 12px;
      margin: 8px 0 4px;
    }

    .nav-item {
      display: flex;
      align-items: center;
      gap: 10px;
      padding: 10px 14px;
      border-radius: 9px;
      cursor: pointer;
      font-size: 13.5px;
      font-weight: 500;
      color: #666;
      border: none;
      background: transparent;
      width: 100%;
      text-align: left;
      transition: background 0.15s, color 0.15s;
    }

    .nav-item .icon { font-size: 17px; line-height: 1; }

    .nav-item:hover { background: #1c1c28; color: #bbb; }

    .nav-item.active {
      background: linear-gradient(135deg, rgba(108,99,255,.25), rgba(168,85,247,.2));
      color: #fff;
      border: 1px solid rgba(108,99,255,.3);
    }

    .nav-item.active .icon { filter: drop-shadow(0 0 4px #a855f7); }

    .sidebar-footer {
      margin-top: auto;
      padding: 0 8px;
    }

    .badge {
      display: flex;
      align-items: center;
      gap: 6px;
      background: #12201a;
      border: 1px solid #1a4a2e;
      color: #4ade80;
      font-size: 10.5px;
      padding: 7px 12px;
      border-radius: 8px;
    }

    .badge::before {
      content: '';
      width: 6px; height: 6px;
      background: #4ade80;
      border-radius: 50%;
      flex-shrink: 0;
    }

    /* ── Main content ── */
    .main {
      flex: 1;
      display: flex;
      flex-direction: column;
      overflow-y: auto;
      background: #0f0f13;
    }

    .topbar {
      padding: 28px 48px 0;
      border-bottom: 1px solid #1a1a24;
      padding-bottom: 20px;
    }

    .topbar h2 { font-size: 22px; font-weight: 700; color: #fff; }
    .topbar p  { font-size: 13px; color: #555; margin-top: 4px; }

    .content {
      padding: 36px 48px;
      flex: 1;
    }

    /* ── Panes ── */
    .pane { display: none; max-width: 580px; }
    .pane.active { display: block; }

    /* ── Drop zone ── */
    label {
      display: block;
      font-size: 13px;
      color: #888;
      margin-bottom: 7px;
      font-weight: 500;
    }

    .drop-zone {
      border: 2px dashed #2a2a3a;
      border-radius: 12px;
      padding: 36px 24px;
      text-align: center;
      cursor: pointer;
      transition: border-color 0.2s, background 0.2s;
      margin-bottom: 20px;
      position: relative;
    }

    .drop-zone:hover, .drop-zone.dragover {
      border-color: #6c63ff;
      background: #13131e;
    }

    .drop-zone input[type="file"] {
      position: absolute; inset: 0;
      opacity: 0; cursor: pointer;
      width: 100%; height: 100%;
    }

    .drop-icon { font-size: 32px; margin-bottom: 10px; }
    .drop-zone .hint { font-size: 13px; color: #555; }
    .drop-zone .filename {
      font-size: 13px; color: #a78bfa; font-weight: 500;
      margin-top: 8px; word-break: break-all;
    }

    /* ── File list (merge) ── */
    .file-list { list-style: none; margin-bottom: 20px; display: flex; flex-direction: column; gap: 6px; }

    .file-list li {
      display: flex; align-items: center; justify-content: space-between;
      background: #111118; border: 1px solid #2a2a3a;
      border-radius: 8px; padding: 9px 14px;
      font-size: 13px; color: #a78bfa;
    }

    .file-list li .rm {
      background: none; border: none; color: #444;
      cursor: pointer; font-size: 16px; line-height: 1; padding: 0 2px;
    }
    .file-list li .rm:hover { color: #f87171; }

    /* ── Inputs ── */
    .input-wrap { position: relative; margin-bottom: 20px; }

    .input-wrap input[type="password"],
    .input-wrap input[type="text"],
    .input-wrap input[type="number"] {
      width: 100%;
      background: #111118;
      border: 1px solid #2a2a3a;
      border-radius: 9px;
      padding: 12px 44px 12px 15px;
      font-size: 14px;
      color: #e0e0e0;
      outline: none;
      transition: border-color 0.2s;
    }

    .input-wrap input[type="number"] { padding-right: 15px; }
    .input-wrap input:focus { border-color: #6c63ff; }

    .toggle-eye {
      position: absolute; right: 13px; top: 50%; transform: translateY(-50%);
      background: none; border: none; cursor: pointer;
      color: #555; font-size: 16px; padding: 2px; line-height: 1;
    }
    .toggle-eye:hover { color: #999; }

    .input-row { display: flex; gap: 14px; margin-bottom: 20px; }
    .input-row .input-wrap { margin-bottom: 0; flex: 1; }

    /* ── Buttons ── */
    .btn {
      width: 100%; padding: 13px;
      background: linear-gradient(135deg, #6c63ff, #a855f7);
      border: none; border-radius: 9px;
      color: #fff; font-size: 15px; font-weight: 600;
      cursor: pointer; transition: opacity 0.2s, transform 0.1s;
      letter-spacing: .3px;
    }
    .btn:hover:not(:disabled) { opacity: 0.88; }
    .btn:active:not(:disabled) { transform: scale(0.99); }
    .btn:disabled { opacity: 0.4; cursor: not-allowed; }

    /* ── Status ── */
    .status {
      margin-top: 18px; padding: 13px 16px;
      border-radius: 9px; font-size: 13px; display: none;
    }
    .status.error   { background: #2a1010; border: 1px solid #5a1a1a; color: #f87171; display: block; }
    .status.success { background: #0f2a1a; border: 1px solid #1a5a2e; color: #4ade80; display: block; }
    .status.loading {
      background: #1a1a2e; border: 1px solid #2a2a5a; color: #818cf8;
      display: flex; align-items: center; gap: 10px;
    }

    .spinner {
      width: 16px; height: 16px;
      border: 2px solid #3a3a6a; border-top-color: #818cf8;
      border-radius: 50%; animation: spin .7s linear infinite; flex-shrink: 0;
    }
    @keyframes spin { to { transform: rotate(360deg); } }

    .hint-text { font-size: 12px; color: #444; margin-top: -14px; margin-bottom: 18px; }
  </style>
</head>
<body>

  <!-- Sidebar -->
  <aside class="sidebar">
    <div class="brand">
      <div class="brand-icon">📑</div>
      <div>
        <h1>PDF Tools</h1>
        <p>All-in-one PDF utility</p>
      </div>
    </div>

    <span class="nav-label">Tools</span>

    <button class="nav-item active" onclick="switchTab('unlock','Unlock PDF','Remove password protection from a PDF')">
      <span class="icon">🔓</span> Unlock
    </button>
    <button class="nav-item" onclick="switchTab('lock','Lock PDF','Add password protection to a PDF')">
      <span class="icon">🔒</span> Lock
    </button>
    <button class="nav-item" onclick="switchTab('merge','Merge PDFs','Combine multiple PDFs into one file')">
      <span class="icon">🗂</span> Merge
    </button>
    <button class="nav-item" onclick="switchTab('split','Split PDF','Split a PDF into individual pages')">
      <span class="icon">✂️</span> Split
    </button>
    <button class="nav-item" onclick="switchTab('toxlsx','PDF to XLSX','Extract tables &amp; text from PDF into Excel')">
      <span class="icon">📊</span> To XLSX
    </button>
    <button class="nav-item" onclick="switchTab('compress','Compress PDF','Reduce PDF file size by compressing streams')">
      <span class="icon">🗜</span> Compress
    </button>

    <div class="sidebar-footer">
      <div class="badge">100% Local — files never leave your machine</div>
    </div>
  </aside>

  <!-- Main -->
  <main class="main">
    <div class="topbar">
      <h2 id="topTitle">Unlock PDF</h2>
      <p id="topDesc">Remove password protection from a PDF</p>
    </div>

    <div class="content">

      <!-- UNLOCK -->
      <div class="pane active" id="pane-unlock">
        <form id="unlockForm">
          <label>Select PDF File</label>
          <div class="drop-zone" id="unlockDrop">
            <input type="file" id="unlockFile" accept=".pdf"/>
            <div class="drop-icon">📄</div>
            <div class="hint">Click to browse or drag &amp; drop</div>
            <div class="filename" id="unlockFileName"></div>
          </div>
          <label>PDF Password</label>
          <div class="input-wrap">
            <input type="password" id="unlockPwd" placeholder="Enter the PDF password" autocomplete="off"/>
            <button type="button" class="toggle-eye" onclick="togglePwd('unlockPwd')">👁</button>
          </div>
          <button type="submit" class="btn" id="unlockBtn" disabled>Unlock &amp; Download</button>
        </form>
        <div class="status" id="unlockStatus"></div>
      </div>

      <!-- LOCK -->
      <div class="pane" id="pane-lock">
        <form id="lockForm">
          <label>Select PDF File</label>
          <div class="drop-zone" id="lockDrop">
            <input type="file" id="lockFile" accept=".pdf"/>
            <div class="drop-icon">📄</div>
            <div class="hint">Click to browse or drag &amp; drop</div>
            <div class="filename" id="lockFileName"></div>
          </div>
          <label>New Password</label>
          <div class="input-wrap">
            <input type="password" id="lockPwd" placeholder="Set a password for the PDF" autocomplete="new-password"/>
            <button type="button" class="toggle-eye" onclick="togglePwd('lockPwd')">👁</button>
          </div>
          <button type="submit" class="btn" id="lockBtn" disabled>Lock &amp; Download</button>
        </form>
        <div class="status" id="lockStatus"></div>
      </div>

      <!-- MERGE -->
      <div class="pane" id="pane-merge">
        <form id="mergeForm">
          <label>Add PDF Files (in order)</label>
          <div class="drop-zone" id="mergeDrop" style="margin-bottom:12px;">
            <input type="file" id="mergeFile" accept=".pdf" multiple/>
            <div class="drop-icon">📂</div>
            <div class="hint">Click to browse or drag &amp; drop (multiple OK)</div>
          </div>
          <ul class="file-list" id="mergeList"></ul>
          <button type="submit" class="btn" id="mergeBtn" disabled>Merge &amp; Download</button>
        </form>
        <div class="status" id="mergeStatus"></div>
      </div>

      <!-- SPLIT -->
      <div class="pane" id="pane-split">
        <form id="splitForm">
          <label>Select PDF File</label>
          <div class="drop-zone" id="splitDrop">
            <input type="file" id="splitFile" accept=".pdf"/>
            <div class="drop-icon">📄</div>
            <div class="hint">Click to browse or drag &amp; drop</div>
            <div class="filename" id="splitFileName"></div>
          </div>
          <label>Page Range <span style="color:#444;font-weight:400">(optional — leave blank to split all)</span></label>
          <div class="input-row">
            <div class="input-wrap">
              <label style="font-size:12px;margin-bottom:5px;">From page</label>
              <input type="number" id="splitFrom" placeholder="1" min="1"/>
            </div>
            <div class="input-wrap">
              <label style="font-size:12px;margin-bottom:5px;">To page</label>
              <input type="number" id="splitTo" placeholder="end" min="1"/>
            </div>
          </div>
          <p class="hint-text">Each page is saved as a separate PDF inside a ZIP archive.</p>
          <button type="submit" class="btn" id="splitBtn" disabled>Split &amp; Download</button>
        </form>
        <div class="status" id="splitStatus"></div>
      </div>

      <!-- TO XLSX -->
      <div class="pane" id="pane-toxlsx">
        <form id="toxlsxForm">
          <label>Select PDF File</label>
          <div class="drop-zone" id="toxlsxDrop">
            <input type="file" id="toxlsxFile" accept=".pdf"/>
            <div class="drop-icon">📄</div>
            <div class="hint">Click to browse or drag &amp; drop</div>
            <div class="filename" id="toxlsxFileName"></div>
          </div>
          <p class="hint-text">Tables &amp; text are extracted page-by-page into separate sheets.</p>
          <button type="submit" class="btn" id="toxlsxBtn" disabled>Convert &amp; Download</button>
        </form>
        <div class="status" id="toxlsxStatus"></div>
      </div>

      <!-- COMPRESS -->
      <div class="pane" id="pane-compress">
        <form id="compressForm">
          <label>Select PDF File</label>
          <div class="drop-zone" id="compressDrop">
            <input type="file" id="compressFile" accept=".pdf"/>
            <div class="drop-icon">📄</div>
            <div class="hint">Click to browse or drag &amp; drop</div>
            <div class="filename" id="compressFileName"></div>
          </div>
          <p class="hint-text">Removes redundant data and compresses streams to reduce file size.</p>
          <button type="submit" class="btn" id="compressBtn" disabled>Compress &amp; Download</button>
        </form>
        <div class="status" id="compressStatus"></div>
      </div>

    </div><!-- /content -->
  </main>

  <script>
    const TABS = ['unlock','lock','merge','split','toxlsx','compress'];

    function switchTab(name, title, desc) {
      TABS.forEach((t, i) => {
        document.querySelectorAll('.nav-item')[i].classList.toggle('active', t === name);
        document.getElementById('pane-' + t).classList.toggle('active', t === name);
      });
      document.getElementById('topTitle').textContent = title;
      document.getElementById('topDesc').innerHTML    = desc;
    }

    function togglePwd(id) {
      const el = document.getElementById(id);
      el.type = el.type === 'password' ? 'text' : 'password';
    }

    function setStatus(id, type, msg) {
      const el = document.getElementById(id);
      el.className = 'status ' + type;
      el.innerHTML = type === 'loading'
        ? '<div class="spinner"></div><span>' + msg + '</span>'
        : msg;
    }

    function triggerDownload(blob, filename) {
      const url = URL.createObjectURL(blob);
      const a   = document.createElement('a');
      a.href = url; a.download = filename; a.click();
      URL.revokeObjectURL(url);
    }

    function setupDrop(dropId, inputId, multi) {
      const drop  = document.getElementById(dropId);
      const input = document.getElementById(inputId);
      drop.addEventListener('dragover', e => { e.preventDefault(); drop.classList.add('dragover'); });
      drop.addEventListener('dragleave', () => drop.classList.remove('dragover'));
      drop.addEventListener('drop', e => {
        e.preventDefault(); drop.classList.remove('dragover');
        const files = [...e.dataTransfer.files].filter(f => f.name.toLowerCase().endsWith('.pdf'));
        if (!files.length) return;
        if (multi) { addMergeFiles(files); return; }
        const dt = new DataTransfer(); dt.items.add(files[0]); input.files = dt.files;
        input.dispatchEvent(new Event('change'));
      });
    }

    /* ── UNLOCK ── */
    setupDrop('unlockDrop','unlockFile', false);
    const unlockFile = document.getElementById('unlockFile');
    const unlockPwd  = document.getElementById('unlockPwd');
    const unlockBtn  = document.getElementById('unlockBtn');
    const checkUnlock = () => { unlockBtn.disabled = !(unlockFile.files.length && unlockPwd.value); };
    unlockFile.addEventListener('change', () => {
      document.getElementById('unlockFileName').textContent = unlockFile.files[0]?.name || '';
      checkUnlock();
    });
    unlockPwd.addEventListener('input', checkUnlock);
    document.getElementById('unlockForm').addEventListener('submit', async e => {
      e.preventDefault();
      unlockBtn.disabled = true;
      setStatus('unlockStatus','loading','Unlocking PDF...');
      const fd = new FormData();
      fd.append('pdf', unlockFile.files[0]);
      fd.append('password', unlockPwd.value);
      try {
        const res = await fetch('/unlock', { method:'POST', body:fd });
        if (res.ok) { triggerDownload(await res.blob(), unlockFile.files[0].name.replace(/\.pdf$/i,'')+'_unlocked.pdf'); setStatus('unlockStatus','success','✓ Unlocked! Download started.'); }
        else { const e = await res.json(); setStatus('unlockStatus','error','✗ '+(e.error||'Error')); }
      } catch { setStatus('unlockStatus','error','✗ Could not reach server.'); }
      finally { unlockBtn.disabled = false; checkUnlock(); }
    });

    /* ── LOCK ── */
    setupDrop('lockDrop','lockFile', false);
    const lockFile = document.getElementById('lockFile');
    const lockPwd  = document.getElementById('lockPwd');
    const lockBtn  = document.getElementById('lockBtn');
    const checkLock = () => { lockBtn.disabled = !(lockFile.files.length && lockPwd.value); };
    lockFile.addEventListener('change', () => {
      document.getElementById('lockFileName').textContent = lockFile.files[0]?.name || '';
      checkLock();
    });
    lockPwd.addEventListener('input', checkLock);
    document.getElementById('lockForm').addEventListener('submit', async e => {
      e.preventDefault();
      lockBtn.disabled = true;
      setStatus('lockStatus','loading','Locking PDF...');
      const fd = new FormData();
      fd.append('pdf', lockFile.files[0]);
      fd.append('password', lockPwd.value);
      try {
        const res = await fetch('/lock', { method:'POST', body:fd });
        if (res.ok) { triggerDownload(await res.blob(), lockFile.files[0].name.replace(/\.pdf$/i,'')+'_locked.pdf'); setStatus('lockStatus','success','✓ Locked! Download started.'); }
        else { const e = await res.json(); setStatus('lockStatus','error','✗ '+(e.error||'Error')); }
      } catch { setStatus('lockStatus','error','✗ Could not reach server.'); }
      finally { lockBtn.disabled = false; checkLock(); }
    });

    /* ── MERGE ── */
    setupDrop('mergeDrop','mergeFile', true);
    const mergeFile = document.getElementById('mergeFile');
    const mergeBtn  = document.getElementById('mergeBtn');
    const mergeList = document.getElementById('mergeList');
    let mergeFiles  = [];
    function renderMergeList() {
      mergeList.innerHTML = '';
      mergeFiles.forEach((f, i) => {
        const li = document.createElement('li');
        li.innerHTML = `<span>📄 ${f.name}</span><button type="button" class="rm" data-i="${i}">✕</button>`;
        mergeList.appendChild(li);
      });
      mergeList.querySelectorAll('.rm').forEach(b => b.addEventListener('click', () => {
        mergeFiles.splice(+b.dataset.i, 1); renderMergeList();
      }));
      mergeBtn.disabled = mergeFiles.length < 2;
    }
    function addMergeFiles(files) { mergeFiles.push(...files); renderMergeList(); }
    mergeFile.addEventListener('change', () => { addMergeFiles([...mergeFile.files]); mergeFile.value=''; });
    document.getElementById('mergeForm').addEventListener('submit', async e => {
      e.preventDefault();
      mergeBtn.disabled = true;
      setStatus('mergeStatus','loading','Merging PDFs...');
      const fd = new FormData();
      mergeFiles.forEach(f => fd.append('pdfs', f));
      try {
        const res = await fetch('/merge', { method:'POST', body:fd });
        if (res.ok) { triggerDownload(await res.blob(), 'merged.pdf'); setStatus('mergeStatus','success','✓ Merged! Download started.'); }
        else { const e = await res.json(); setStatus('mergeStatus','error','✗ '+(e.error||'Error')); }
      } catch { setStatus('mergeStatus','error','✗ Could not reach server.'); }
      finally { mergeBtn.disabled = mergeFiles.length < 2; }
    });

    /* ── SPLIT ── */
    setupDrop('splitDrop','splitFile', false);
    const splitFile = document.getElementById('splitFile');
    const splitBtn  = document.getElementById('splitBtn');
    splitFile.addEventListener('change', () => {
      document.getElementById('splitFileName').textContent = splitFile.files[0]?.name || '';
      splitBtn.disabled = !splitFile.files.length;
    });
    document.getElementById('splitForm').addEventListener('submit', async e => {
      e.preventDefault();
      splitBtn.disabled = true;
      setStatus('splitStatus','loading','Splitting PDF...');
      const fd = new FormData();
      fd.append('pdf', splitFile.files[0]);
      const fv = document.getElementById('splitFrom').value.trim();
      const tv = document.getElementById('splitTo').value.trim();
      if (fv) fd.append('from_page', fv);
      if (tv) fd.append('to_page', tv);
      try {
        const res = await fetch('/split', { method:'POST', body:fd });
        if (res.ok) { triggerDownload(await res.blob(), splitFile.files[0].name.replace(/\.pdf$/i,'')+'_split.zip'); setStatus('splitStatus','success','✓ Split! ZIP download started.'); }
        else { const e = await res.json(); setStatus('splitStatus','error','✗ '+(e.error||'Error')); }
      } catch { setStatus('splitStatus','error','✗ Could not reach server.'); }
      finally { splitBtn.disabled = false; }
    });

    /* ── TO XLSX ── */
    setupDrop('toxlsxDrop','toxlsxFile', false);
    const toxlsxFile = document.getElementById('toxlsxFile');
    const toxlsxBtn  = document.getElementById('toxlsxBtn');
    toxlsxFile.addEventListener('change', () => {
      document.getElementById('toxlsxFileName').textContent = toxlsxFile.files[0]?.name || '';
      toxlsxBtn.disabled = !toxlsxFile.files.length;
    });
    document.getElementById('toxlsxForm').addEventListener('submit', async e => {
      e.preventDefault();
      toxlsxBtn.disabled = true;
      setStatus('toxlsxStatus','loading','Converting to XLSX...');
      const fd = new FormData();
      fd.append('pdf', toxlsxFile.files[0]);
      try {
        const res = await fetch('/to_xlsx', { method:'POST', body:fd });
        if (res.ok) { triggerDownload(await res.blob(), toxlsxFile.files[0].name.replace(/\.pdf$/i,'')+'.xlsx'); setStatus('toxlsxStatus','success','✓ Converted! Download started.'); }
        else { const e = await res.json(); setStatus('toxlsxStatus','error','✗ '+(e.error||'Error')); }
      } catch { setStatus('toxlsxStatus','error','✗ Could not reach server.'); }
      finally { toxlsxBtn.disabled = false; }
    });

    /* ── COMPRESS ── */
    setupDrop('compressDrop','compressFile', false);
    const compressFile = document.getElementById('compressFile');
    const compressBtn  = document.getElementById('compressBtn');
    compressFile.addEventListener('change', () => {
      document.getElementById('compressFileName').textContent = compressFile.files[0]?.name || '';
      compressBtn.disabled = !compressFile.files.length;
    });
    document.getElementById('compressForm').addEventListener('submit', async e => {
      e.preventDefault();
      compressBtn.disabled = true;
      setStatus('compressStatus','loading','Compressing PDF...');
      const fd = new FormData();
      fd.append('pdf', compressFile.files[0]);
      try {
        const res = await fetch('/compress', { method:'POST', body:fd });
        if (res.ok) { triggerDownload(await res.blob(), compressFile.files[0].name.replace(/\.pdf$/i,'')+'_compressed.pdf'); setStatus('compressStatus','success','✓ Compressed! Download started.'); }
        else { const e = await res.json(); setStatus('compressStatus','error','✗ '+(e.error||'Error')); }
      } catch { setStatus('compressStatus','error','✗ Could not reach server.'); }
      finally { compressBtn.disabled = false; }
    });
  </script>
</body>
</html>
"""


@app.route("/")
def index():
    return render_template_string(HTML)


# ── Unlock ───────────────────────────────────────────────────────────────────

@app.route("/unlock", methods=["POST"])
def unlock():
    if "pdf" not in request.files:
        return jsonify({"error": "No file provided."}), 400
    pdf_file = request.files["pdf"]
    password = request.form.get("password", "")
    if not pdf_file.filename.lower().endswith(".pdf"):
        return jsonify({"error": "Please upload a valid PDF file."}), 400
    try:
        with pikepdf.open(io.BytesIO(pdf_file.read()), password=password) as pdf:
            output = io.BytesIO()
            pdf.save(output)
            output.seek(0)
        base = os.path.splitext(pdf_file.filename)[0]
        return send_file(output, mimetype="application/pdf", as_attachment=True,
                         download_name=f"{base}_unlocked.pdf")
    except pikepdf.PasswordError:
        return jsonify({"error": "Incorrect password. Please try again."}), 401
    except pikepdf.PdfError as e:
        return jsonify({"error": f"Could not process PDF: {e}"}), 422
    except Exception as e:
        return jsonify({"error": f"Unexpected error: {e}"}), 500


# ── Lock ─────────────────────────────────────────────────────────────────────

@app.route("/lock", methods=["POST"])
def lock():
    if "pdf" not in request.files:
        return jsonify({"error": "No file provided."}), 400
    pdf_file = request.files["pdf"]
    password = request.form.get("password", "")
    if not pdf_file.filename.lower().endswith(".pdf"):
        return jsonify({"error": "Please upload a valid PDF file."}), 400
    if not password:
        return jsonify({"error": "A password is required to lock the PDF."}), 400
    try:
        with pikepdf.open(io.BytesIO(pdf_file.read())) as pdf:
            output = io.BytesIO()
            pdf.save(output, encryption=pikepdf.Encryption(owner=password, user=password, R=6))
            output.seek(0)
        base = os.path.splitext(pdf_file.filename)[0]
        return send_file(output, mimetype="application/pdf", as_attachment=True,
                         download_name=f"{base}_locked.pdf")
    except pikepdf.PdfError as e:
        return jsonify({"error": f"Could not process PDF: {e}"}), 422
    except Exception as e:
        return jsonify({"error": f"Unexpected error: {e}"}), 500


# ── Merge ────────────────────────────────────────────────────────────────────

@app.route("/merge", methods=["POST"])
def merge():
    files = request.files.getlist("pdfs")
    if len(files) < 2:
        return jsonify({"error": "Please provide at least 2 PDF files to merge."}), 400
    try:
        merged = pikepdf.Pdf.new()
        for f in files:
            if not f.filename.lower().endswith(".pdf"):
                return jsonify({"error": f"'{f.filename}' is not a PDF."}), 400
            with pikepdf.open(io.BytesIO(f.read())) as src:
                merged.pages.extend(src.pages)
        output = io.BytesIO()
        merged.save(output)
        merged.close()
        output.seek(0)
        return send_file(output, mimetype="application/pdf", as_attachment=True,
                         download_name="merged.pdf")
    except pikepdf.PdfError as e:
        return jsonify({"error": f"Could not process PDF: {e}"}), 422
    except Exception as e:
        return jsonify({"error": f"Unexpected error: {e}"}), 500


# ── Split ────────────────────────────────────────────────────────────────────

@app.route("/split", methods=["POST"])
def split():
    if "pdf" not in request.files:
        return jsonify({"error": "No file provided."}), 400
    pdf_file = request.files["pdf"]
    if not pdf_file.filename.lower().endswith(".pdf"):
        return jsonify({"error": "Please upload a valid PDF file."}), 400
    try:
        from_page = int(request.form.get("from_page", 1))
        to_page   = request.form.get("to_page", None)
        with pikepdf.open(io.BytesIO(pdf_file.read())) as src:
            total   = len(src.pages)
            to_page = int(to_page) if to_page else total
            from_page = max(1, min(from_page, total))
            to_page   = max(from_page, min(to_page, total))
            base    = os.path.splitext(pdf_file.filename)[0]
            zip_buf = io.BytesIO()
            with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                for i in range(from_page - 1, to_page):
                    pg = pikepdf.Pdf.new()
                    pg.pages.append(src.pages[i])
                    pb = io.BytesIO()
                    pg.save(pb); pg.close(); pb.seek(0)
                    zf.writestr(f"{base}_page_{i+1}.pdf", pb.read())
        zip_buf.seek(0)
        return send_file(zip_buf, mimetype="application/zip", as_attachment=True,
                         download_name=f"{base}_split.zip")
    except ValueError:
        return jsonify({"error": "Page numbers must be integers."}), 400
    except pikepdf.PdfError as e:
        return jsonify({"error": f"Could not process PDF: {e}"}), 422
    except Exception as e:
        return jsonify({"error": f"Unexpected error: {e}"}), 500


# ── To XLSX ──────────────────────────────────────────────────────────────────

@app.route("/to_xlsx", methods=["POST"])
def to_xlsx():
    if "pdf" not in request.files:
        return jsonify({"error": "No file provided."}), 400
    pdf_file = request.files["pdf"]
    if not pdf_file.filename.lower().endswith(".pdf"):
        return jsonify({"error": "Please upload a valid PDF file."}), 400
    try:
        pdf_bytes = pdf_file.read()
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page_num, page in enumerate(pdf.pages, start=1):
                ws  = wb.create_sheet(title=f"Page {page_num}")
                row = 1
                tables = page.extract_tables()
                for table in tables:
                    for tr in table:
                        for ci, cell in enumerate(tr, start=1):
                            ws.cell(row=row, column=ci, value=cell or "")
                        row += 1
                    row += 1
                if not tables:
                    for line in (page.extract_text() or "").splitlines():
                        ws.cell(row=row, column=1, value=line)
                        row += 1
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        base = os.path.splitext(pdf_file.filename)[0]
        return send_file(output,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=f"{base}.xlsx")
    except Exception as e:
        return jsonify({"error": f"Unexpected error: {e}"}), 500


# ── Compress ─────────────────────────────────────────────────────────────────

@app.route("/compress", methods=["POST"])
def compress():
    if "pdf" not in request.files:
        return jsonify({"error": "No file provided."}), 400
    pdf_file = request.files["pdf"]
    if not pdf_file.filename.lower().endswith(".pdf"):
        return jsonify({"error": "Please upload a valid PDF file."}), 400
    try:
        pdf_bytes = pdf_file.read()
        with pikepdf.open(io.BytesIO(pdf_bytes)) as pdf:
            output = io.BytesIO()
            pdf.save(output, compress_streams=True,
                     stream_decode_level=pikepdf.StreamDecodeLevel.generalized,
                     object_stream_mode=pikepdf.ObjectStreamMode.generate,
                     recompress_flate=True)
            output.seek(0)
        orig = len(pdf_bytes)
        comp = output.getbuffer().nbytes
        print(f"[Compress] {orig} -> {comp} bytes ({max(0, 100 - comp*100//orig)}% saved)")
        base = os.path.splitext(pdf_file.filename)[0]
        return send_file(output, mimetype="application/pdf", as_attachment=True,
                         download_name=f"{base}_compressed.pdf")
    except pikepdf.PdfError as e:
        return jsonify({"error": f"Could not process PDF: {e}"}), 422
    except Exception as e:
        return jsonify({"error": f"Unexpected error: {e}"}), 500


# ── Entry point ───────────────────────────────────────────────────────────────

def open_browser():
    webbrowser.open("http://127.0.0.1:5000")

if __name__ == "__main__":
    threading.Timer(1.0, open_browser).start()
    print("=" * 50)
    print("  PDF Tools running at http://127.0.0.1:5000")
    print("  Press Ctrl+C to stop.")
    print("=" * 50)
    app.run(debug=False, port=5000)
